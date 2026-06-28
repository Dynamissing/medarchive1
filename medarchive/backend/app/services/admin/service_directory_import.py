from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.db.models import Service, ServiceSynonym

logger = get_logger(__name__)

ORGANIZER_COLUMNS = {
    "source_row_id": ("ID", "id", "Id"),
    "specialty": ("Специальность", "specialty", "category", "Category"),
    "code": ("Code", "code"),
    "name_ru": ("Name_ru", "name_ru", "Name", "name", "Название"),
    "tariff_code": ("TarificatrCode", "TarificatorCode", "tariff_code", "tarificatr_code"),
}

SPACE_RE = re.compile(r"\s+")
PAREN_RE = re.compile(r"\([^)]*\)")


@dataclass(frozen=True)
class ServiceImportRow:
    source_type: str
    source_row_number: int | None
    source_row_id: str | None
    code: str | None
    tariff_code: str | None
    category: str | None
    specialty: str | None
    name_ru: str
    normalized_name: str
    normalized_specialty: str | None
    warnings: list[str] = field(default_factory=list)
    raw_data: dict[str, str | None] = field(default_factory=dict)

    @property
    def source_hash(self) -> str:
        payload = {
            "source_row_id": self.source_row_id,
            "code": self.code,
            "tariff_code": self.tariff_code,
            "category": self.category,
            "specialty": self.specialty,
            "name_ru": self.name_ru,
        }
        encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()


@dataclass(frozen=True)
class ServiceImportResult:
    batch: str
    source_path: Path
    rows_seen: int
    imported: int
    updated: int
    skipped: int
    warnings: list[str]


def normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    cleaned = "".join(
        character if character.isalnum() or character.isspace() else " "
        for character in value.casefold()
    )
    return SPACE_RE.sub(" ", cleaned).strip()


def generate_synonyms(name: str, specialty: str | None = None) -> list[str]:
    candidates: list[str] = [name]
    no_parenthetical = PAREN_RE.sub(" ", name)
    if no_parenthetical != name:
        candidates.append(no_parenthetical)

    normalized_name = normalize_text(name)
    if normalized_name:
        candidates.append(normalized_name)

    if specialty:
        combined = f"{name} {specialty}"
        candidates.append(combined)

    seen: set[str] = set()
    synonyms: list[str] = []
    for candidate in candidates:
        normalized = normalize_text(candidate)
        if normalized and normalized not in seen:
            seen.add(normalized)
            synonyms.append(candidate.strip())
    return synonyms


def import_service_directory(db: Session, source_path: Path, batch: str | None = None) -> ServiceImportResult:
    resolved_path = source_path.resolve()
    import_batch = batch or build_default_batch(resolved_path)
    rows = read_service_rows(resolved_path)

    imported = 0
    updated = 0
    skipped = 0
    import_warnings: list[str] = []

    for row in rows:
        if row.warnings:
            import_warnings.extend(format_row_warning(row, warning) for warning in row.warnings)

        if not row.name_ru:
            skipped += 1
            continue

        existing = db.scalar(
            select(Service).where(
                Service.import_batch == import_batch,
                Service.source_hash == row.source_hash,
            )
        )
        if existing is None:
            service = Service(
                import_batch=import_batch,
                source_type=row.source_type,
                source_hash=row.source_hash,
                source_row_number=row.source_row_number,
                source_row_id=row.source_row_id,
                code=row.code,
                tariff_code=row.tariff_code,
                category=row.category,
                specialty=row.specialty,
                name_ru=row.name_ru,
                normalized_name=row.normalized_name,
                normalized_specialty=row.normalized_specialty,
                warnings=row.warnings,
                raw_data=row.raw_data,
            )
            db.add(service)
            db.flush()
            imported += 1
        else:
            service = existing
            service.source_row_number = row.source_row_number
            service.source_row_id = row.source_row_id
            service.code = row.code
            service.tariff_code = row.tariff_code
            service.category = row.category
            service.specialty = row.specialty
            service.name_ru = row.name_ru
            service.normalized_name = row.normalized_name
            service.normalized_specialty = row.normalized_specialty
            service.warnings = row.warnings
            service.raw_data = row.raw_data
            service.synonyms.clear()
            db.flush()
            updated += 1

        for synonym in generate_synonyms(row.name_ru, row.specialty):
            db.add(
                ServiceSynonym(
                    service=service,
                    value=synonym,
                    normalized_value=normalize_text(synonym),
                    source="deterministic",
                )
            )

    db.commit()

    for warning in import_warnings:
        logger.warning(warning)

    return ServiceImportResult(
        batch=import_batch,
        source_path=resolved_path,
        rows_seen=len(rows),
        imported=imported,
        updated=updated,
        skipped=skipped,
        warnings=import_warnings,
    )


def build_default_batch(source_path: Path) -> str:
    stat = source_path.stat()
    payload = f"{source_path.name}:{stat.st_size}:{int(stat.st_mtime)}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def read_service_rows(source_path: Path) -> list[ServiceImportRow]:
    suffix = source_path.suffix.casefold()
    if suffix == ".xlsx":
        return read_xlsx_rows(source_path)
    if suffix == ".json":
        return read_json_rows(source_path)
    raise ValueError(f"Unsupported service directory format: {source_path.suffix}")


def read_json_rows(source_path: Path) -> list[ServiceImportRow]:
    payload = json.loads(source_path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        raw_rows = payload.get("services", payload.get("rows", []))
    else:
        raw_rows = payload
    if not isinstance(raw_rows, list):
        raise ValueError("JSON service directory must contain a list of rows")

    rows: list[ServiceImportRow] = []
    for index, raw_row in enumerate(raw_rows, start=1):
        if not isinstance(raw_row, dict):
            rows.append(
                build_import_row(
                    source_type="json",
                    source_row_number=index,
                    raw_data={},
                    warnings=["JSON row is not an object"],
                )
            )
            continue
        rows.append(
            build_import_row(
                source_type="json",
                source_row_number=index,
                raw_data={str(key): stringify_value(value) for key, value in raw_row.items()},
            )
        )
    return rows


def read_xlsx_rows(source_path: Path) -> list[ServiceImportRow]:
    workbook = load_workbook(source_path, data_only=True, read_only=False)
    worksheet = workbook.active
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1), ())
    headers = [stringify_value(cell.value) or "" for cell in header_cells]

    rows: list[ServiceImportRow] = []
    for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
        raw_data: dict[str, str | None] = {}
        warnings: list[str] = []
        for header, cell in zip(headers, cells, strict=False):
            if not header:
                continue
            if cell.data_type == "e":
                raw_data[header] = None
                warnings.append(f"Cell {cell.coordinate} contains spreadsheet error {cell.value!r}")
            else:
                raw_data[header] = stringify_value(cell.value)
        rows.append(
            build_import_row(
                source_type="xlsx",
                source_row_number=row_number,
                raw_data=raw_data,
                warnings=warnings,
            )
        )
    return rows


def build_import_row(
    *,
    source_type: str,
    source_row_number: int | None,
    raw_data: dict[str, str | None],
    warnings: list[str] | None = None,
) -> ServiceImportRow:
    row_warnings = list(warnings or [])
    source_row_id = pick_value(raw_data, ORGANIZER_COLUMNS["source_row_id"])
    specialty = pick_value(raw_data, ORGANIZER_COLUMNS["specialty"])
    code = pick_value(raw_data, ORGANIZER_COLUMNS["code"])
    name_ru = pick_value(raw_data, ORGANIZER_COLUMNS["name_ru"])
    tariff_code = pick_value(raw_data, ORGANIZER_COLUMNS["tariff_code"])

    if not name_ru:
        row_warnings.append("Missing required service name")

    normalized_name = normalize_text(name_ru)
    normalized_specialty = normalize_text(specialty) or None

    return ServiceImportRow(
        source_type=source_type,
        source_row_number=source_row_number,
        source_row_id=source_row_id,
        code=code,
        tariff_code=tariff_code,
        category=specialty,
        specialty=specialty,
        name_ru=name_ru or "",
        normalized_name=normalized_name,
        normalized_specialty=normalized_specialty,
        warnings=row_warnings,
        raw_data=raw_data,
    )


def pick_value(raw_data: dict[str, str | None], candidates: tuple[str, ...]) -> str | None:
    for key in candidates:
        value = raw_data.get(key)
        if value:
            return value
    return None


def stringify_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def format_row_warning(row: ServiceImportRow, warning: str) -> str:
    row_label = row.source_row_number if row.source_row_number is not None else "unknown"
    return f"Service import row {row_label}: {warning}"
