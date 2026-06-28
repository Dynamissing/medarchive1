from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.parsed_document import (
    ParsedDocumentResult,
    SpreadsheetPriceVariant,
    SpreadsheetRowCandidate,
)
from app.services.parsers.base import DocumentParser, ParserInput

HEADER_TERMS = (
    "наименование",
    "услуг",
    "услуга",
    "название",
    "name",
    "service",
    "код",
    "code",
    "цена",
    "стоимость",
    "price",
    "тариф",
)
NAME_TERMS = ("наименование", "услуг", "услуга", "название", "name", "service")
PRICE_TERMS = ("цена", "стоимость", "price", "тариф", "сумма")
CODE_TERMS = ("код", "code", "шифр")
SPACE_RE = re.compile(r"\s+")


@dataclass(frozen=True)
class HeaderSelection:
    row_indexes: list[int]
    labels: dict[int, str]


class XlsxParser(DocumentParser):
    parser_name = "xlsx"
    parser_format = "xlsx"

    def parse(self, parser_input: ParserInput) -> ParsedDocumentResult:
        extraction = extract_xlsx_rows(parser_input.source_path)
        return ParsedDocumentResult(
            parser_name=self.parser_name,
            parser_format="xlsx",
            status="parsed",
            source_file_asset_id=parser_input.file_asset_id,
            source_path=str(parser_input.source_path),
            row_candidates=extraction.row_candidates,
            metadata={
                "sheet_count": extraction.sheet_count,
                "candidate_count": len(extraction.row_candidates),
                "mime_type": parser_input.mime_type,
                "extension": parser_input.extension,
            },
            warnings=extraction.warnings,
        )


class XlsParser(DocumentParser):
    parser_name = "xls_via_libreoffice"
    parser_format = "xls"

    def parse(self, parser_input: ParserInput) -> ParsedDocumentResult:
        converted_path: Path | None = None
        warnings: list[str] = []
        try:
            converted_path = convert_xls_to_xlsx(parser_input.source_path)
            extraction = extract_xlsx_rows(converted_path)
            warnings.extend(extraction.warnings)
            warnings.append("XLS converted to XLSX with LibreOffice before parsing.")
            return ParsedDocumentResult(
                parser_name=self.parser_name,
                parser_format="xls",
                status="parsed",
                source_file_asset_id=parser_input.file_asset_id,
                source_path=str(parser_input.source_path),
                row_candidates=extraction.row_candidates,
                metadata={
                    "converted_path": str(converted_path),
                    "sheet_count": extraction.sheet_count,
                    "candidate_count": len(extraction.row_candidates),
                    "mime_type": parser_input.mime_type,
                    "extension": parser_input.extension,
                },
                warnings=warnings,
            )
        finally:
            if converted_path is not None:
                shutil.rmtree(converted_path.parent, ignore_errors=True)


@dataclass(frozen=True)
class SpreadsheetExtraction:
    row_candidates: list[SpreadsheetRowCandidate]
    sheet_count: int
    warnings: list[str]


def extract_xlsx_rows(path: Path) -> SpreadsheetExtraction:
    workbook = load_workbook(path, data_only=True, read_only=False)
    row_candidates: list[SpreadsheetRowCandidate] = []
    warnings: list[str] = []

    for worksheet in workbook.worksheets:
        merged_values = build_merged_value_map(worksheet)
        header = detect_header(worksheet, merged_values)
        if header is None:
            warnings.append(f"No usable header detected in sheet {worksheet.title!r}.")
            continue
        category_path: list[str] = []
        max_header_row = max(header.row_indexes)
        for row_index in range(max_header_row + 1, worksheet.max_row + 1):
            cells = list(worksheet[row_index])
            values = [cell_value(cell, merged_values) for cell in cells]
            if is_empty_row(values):
                continue
            if is_category_row(values, header):
                category = first_text(values)
                if category:
                    category_path = [category]
                continue
            candidate = build_candidate(worksheet.title, row_index, cells, values, header, category_path)
            if candidate is not None:
                row_candidates.append(candidate)

    return SpreadsheetExtraction(
        row_candidates=row_candidates,
        sheet_count=len(workbook.worksheets),
        warnings=warnings,
    )


def detect_header(worksheet: Worksheet, merged_values: dict[str, Any], scan_rows: int = 25) -> HeaderSelection | None:
    best: tuple[int, HeaderSelection] | None = None
    max_row = min(worksheet.max_row, scan_rows)
    for end_row in range(1, max_row + 1):
        for height in (1, 2):
            start_row = end_row - height + 1
            if start_row < 1:
                continue
            selection = build_header_selection(worksheet, merged_values, list(range(start_row, end_row + 1)))
            score = score_header(selection)
            if score <= 0:
                continue
            if best is None or score > best[0]:
                best = (score, selection)
    if best is None or best[0] < 3:
        return None
    return best[1]


def build_header_selection(
    worksheet: Worksheet,
    merged_values: dict[str, Any],
    row_indexes: list[int],
) -> HeaderSelection:
    labels: dict[int, str] = {}
    for column_index in range(1, worksheet.max_column + 1):
        parts: list[str] = []
        for row_index in row_indexes:
            value = cell_value(worksheet.cell(row=row_index, column=column_index), merged_values)
            if value is not None and str(value).strip():
                parts.append(str(value).strip())
        if parts:
            labels[column_index] = normalize_header(" ".join(parts))
    return HeaderSelection(row_indexes=row_indexes, labels=labels)


def score_header(selection: HeaderSelection) -> int:
    labels = list(selection.labels.values())
    joined = " ".join(labels)
    score = sum(1 for term in HEADER_TERMS if term in joined)
    if any(any(term in label for term in NAME_TERMS) for label in labels):
        score += 2
    if any(any(term in label for term in PRICE_TERMS) for label in labels):
        score += 2
    return score


def build_candidate(
    sheet_name: str,
    row_index: int,
    cells: list[Cell | MergedCell],
    raw_values: list[Any],
    header: HeaderSelection,
    category_path: list[str],
) -> SpreadsheetRowCandidate | None:
    values: dict[str, Any] = {}
    source_cells: dict[str, str] = {}
    price_variants: list[SpreadsheetPriceVariant] = []

    for cell in cells:
        label = header.labels.get(cell.column)
        if not label:
            continue
        value = raw_values[cell.column - 1] if cell.column - 1 < len(raw_values) else None
        if is_blank(value):
            continue
        values[label] = value
        source_cells[label] = cell.coordinate
        if is_price_header(label):
            price_variants.append(SpreadsheetPriceVariant(label=label, value=value, cell=cell.coordinate))

    if not values:
        return None
    if not has_item_signal(values, price_variants):
        return None

    return SpreadsheetRowCandidate(
        sheet_name=sheet_name,
        row_index=row_index,
        category_path=list(category_path),
        raw_values=raw_values,
        values=values,
        source_cells=source_cells,
        price_variants=price_variants,
    )


def has_item_signal(values: dict[str, Any], price_variants: list[SpreadsheetPriceVariant]) -> bool:
    has_name = any(any(term in label for term in NAME_TERMS) for label in values)
    has_code = any(any(term in label for term in CODE_TERMS) for label in values)
    return has_name or (has_code and bool(price_variants)) or len(values) >= 2 and bool(price_variants)


def is_category_row(values: list[Any], header: HeaderSelection) -> bool:
    nonblank = [value for value in values if not is_blank(value)]
    if not nonblank or len(nonblank) > 2:
        return False
    text = " ".join(str(value) for value in nonblank)
    if any(term in normalize_header(text) for term in HEADER_TERMS):
        return False
    columns_with_values = [index + 1 for index, value in enumerate(values) if not is_blank(value)]
    if any(is_price_header(header.labels.get(column, "")) for column in columns_with_values):
        return False
    return not any(is_number_like(value) for value in nonblank)


def build_merged_value_map(worksheet: Worksheet) -> dict[str, Any]:
    merged_values: dict[str, Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[worksheet.cell(row=row, column=column).coordinate] = value
    return merged_values


def cell_value(cell: Cell | MergedCell, merged_values: dict[str, Any]) -> Any:
    return merged_values.get(cell.coordinate, cell.value)


def normalize_header(value: str) -> str:
    return SPACE_RE.sub(" ", value.casefold()).strip()


def is_price_header(label: str) -> bool:
    return any(term in label for term in PRICE_TERMS)


def is_empty_row(values: list[Any]) -> bool:
    return all(is_blank(value) for value in values)


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def first_text(values: list[Any]) -> str | None:
    for value in values:
        if not is_blank(value):
            return str(value).strip()
    return None


def is_number_like(value: Any) -> bool:
    if isinstance(value, int | float):
        return True
    try:
        float(str(value).replace(" ", "").replace(",", "."))
    except ValueError:
        return False
    return True


def convert_xls_to_xlsx(source_path: Path) -> Path:
    executable = find_libreoffice()
    if executable is None:
        raise RuntimeError("LibreOffice executable was not found for XLS conversion.")

    output_dir = Path(tempfile.mkdtemp(prefix="medarchive-xls-"))
    command = [
        str(executable),
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(source_path),
    ]
    completed = subprocess.run(command, capture_output=True, text=True, check=False)
    if completed.returncode != 0:
        shutil.rmtree(output_dir, ignore_errors=True)
        raise RuntimeError(f"LibreOffice XLS conversion failed: {completed.stderr or completed.stdout}")

    converted = output_dir / f"{source_path.stem}.xlsx"
    if not converted.exists():
        matches = list(output_dir.glob("*.xlsx"))
        if not matches:
            shutil.rmtree(output_dir, ignore_errors=True)
            raise RuntimeError("LibreOffice XLS conversion did not produce an XLSX file.")
        converted = matches[0]
    return converted


def find_libreoffice() -> Path | None:
    configured = os.getenv("LIBREOFFICE_EXECUTABLE")
    if configured:
        configured_path = Path(configured)
        if configured_path.exists():
            return configured_path
        found_configured = shutil.which(configured)
        if found_configured:
            return Path(found_configured)

    for executable in ("soffice", "libreoffice"):
        found = shutil.which(executable)
        if found:
            return Path(found)
    for candidate in (
        Path("/usr/bin/soffice"),
        Path("/usr/bin/libreoffice"),
        Path("/usr/lib/libreoffice/program/soffice"),
        Path("C:/Program Files/LibreOffice/program/soffice.exe"),
        Path("C:/Program Files (x86)/LibreOffice/program/soffice.exe"),
    ):
        if candidate.exists():
            return candidate
    return None
