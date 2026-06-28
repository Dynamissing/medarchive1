from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.parsers.base import ParserInput
from app.services.parsers.spreadsheet import XlsxParser


def test_xlsx_parser_extracts_candidates_from_messy_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "clinic.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Main"
    sheet["A1"] = "Clinic price list"
    sheet["A2"] = "Generated manually"
    sheet.append([])
    sheet.append(["", "Service", "Prices", None, ""])
    sheet.merge_cells("C4:D4")
    sheet.append(["Code", "Name", "Cash", "Insurance", "Comment"])
    sheet.append(["Diagnostics"])
    sheet.append(["A-1", "MRI brain", 1000, 900, ""])
    sheet.append([])
    sheet.append(["Consultations"])
    sheet.append(["B-1", "Cardiologist consultation", 500, 450, "primary"])

    second = workbook.create_sheet("Second")
    second.append(["note"])
    second.append(["Service name", "Price"])
    second.append(["Blood test", 200])
    workbook.save(workbook_path)

    result = XlsxParser().parse(ParserInput(parser_format="xlsx", source_path=workbook_path))

    assert result.status == "parsed"
    assert result.parser_format == "xlsx"
    assert result.metadata["sheet_count"] == 2
    assert len(result.row_candidates) == 3

    first = result.row_candidates[0]
    assert first.sheet_name == "Main"
    assert first.row_index == 7
    assert first.category_path == ["Diagnostics"]
    assert first.values["code"] == "A-1"
    assert first.values["service name"] == "MRI brain"
    assert [variant.label for variant in first.price_variants] == ["prices cash", "prices insurance"]
    assert [variant.value for variant in first.price_variants] == [1000, 900]
    assert first.source_cells["service name"] == "B7"

    categories = [candidate.category_path for candidate in result.row_candidates]
    assert ["Diagnostics"] in categories
    assert ["Consultations"] in categories
    assert all(candidate.values.get("service name") != "Diagnostics" for candidate in result.row_candidates)


def test_xlsx_parser_skips_sheets_without_header(tmp_path: Path) -> None:
    workbook_path = tmp_path / "no-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Only", "free", "text"])
    sheet.append(["Category"])
    workbook.save(workbook_path)

    result = XlsxParser().parse(ParserInput(parser_format="xlsx", source_path=workbook_path))

    assert result.row_candidates == []
    assert result.warnings == ["No usable header detected in sheet 'Sheet'."]
